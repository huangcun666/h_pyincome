#-*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

from openpyxl import Workbook
from openpyxl import load_workbook
import torndb
import re
import xlrd
from datetime import datetime
from decimal import Decimal
db = torndb.Connection("192.168.2.168", "db_income2","root","deng")
wb =load_workbook(u"/home/hc/relation/hh.xlsx")
sheet1 = wb.get_sheet_by_name('工作表1')   
t_projects_relation=db.query(''' select relation_ids from t_projects_relation ''')
for row,projects_relation in enumerate(t_projects_relation):
    i=1
    for column,item in enumerate(projects_relation.relation_ids.split(',')[:-1]):
        project=db.get('''
        select customer_company from t_projects where id=%s''',item)
        sheet1.cell(row=row+2,column=i).value=item
        if project.customer_company:
            sheet1.cell(row=row+2,column=i+1).value=project.customer_company
        else:
            sheet1.cell(row=row+2,column=i+1).value='无'
        i+=2
wb.save(u"hh.xlsx")

wb =xlrd.open_workbook(u"hh.xlsx")
sheet2 = wb.sheet_by_name('工作表1') 
max_rows=sheet2.nrows-1
for i in range(max_rows):
    row_values=sheet2.row_values(i+1)
    for j in row_values[-1::-2]:
        if j and j!='无':
            for idx,k in enumerate(row_values):
                if k=='无':
                    print(row_values[idx-1])
                    db.execute('''
                    update t_projects set customer_company=%s where id=%s
                    ''',j,row_values[idx-1])
            break


